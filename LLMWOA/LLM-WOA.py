import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import re
import os
import time
from opfunu.cec_based.cec2022 import *
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

# ������������֧��
plt.rcParams["font.family"] = ["SimHei", "WenQuanYi Micro Hei", "Heiti TC"]
plt.rcParams["axes.unicode_minus"] = False  # ��ȷ��ʾ����

# DeepSeekģ�ͷ�װ * deepseek-chat �� deepseek-reasoner ���Ѿ�����Ϊ DeepSeek-V3.1��
#deepseek-chat ��Ӧ DeepSeek-V3.1 �ķ�˼��ģʽ��
#deepseek-reasoner ��Ӧ DeepSeek-V3.1 ��˼��ģʽ.
class DeepSeekModel:
    def __init__(self, api_key: str, model_name: str = "deepseek-reasoner"):
        try:
            from openai import OpenAI
            self.client = OpenAI(
                api_key=api_key,
                base_url="https://api.deepseek.com/v1"
            )
            self.model = model_name
        except ImportError:
            raise ImportError("�밲װopenai��: pip install openai")
        except Exception as e:
            raise Exception(f"��ʼ��DeepSeekģ��ʧ��: {str(e)}")

    def generate_content(self, prompt: str, **kwargs):
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": kwargs.get("system_prompt", "You are a helpful assistant.")},
                    {"role": "user", "content": prompt}
                ],
                temperature=kwargs.get("temperature", 0.7),
                max_tokens=kwargs.get("max_output_tokens", 2048)
            )
            # ������Ӧ����
            class Response:
                def __init__(self, text, raw):
                    self.text = text
                    self.raw = raw
            return Response(response.choices[0].message.content, response)
        except Exception as e:
            print(f"DeepSeek API����ʧ��: {str(e)}")
            # ����һ��Ĭ����Ӧ������������
            class DefaultResponse:
                def __init__(self):
                    self.text = "a=1.0\nb=1.0\n̽������=0.5����������=0.5"
                    self.raw = None
            return DefaultResponse()

# LLM WOA ����
class LLM_WOA_Advisor:
    def __init__(self, llm_model):
        self.llm = llm_model  # LLMģ�ͣ�DeepSeek��
    
    def get_woa_strategy(self, state):
        """���ݵ�ǰ״̬����LLM����WOA�Ĳ����Ͳ��Ե�������"""
        prompt = f"""���Ǿ����Ż��㷨��WOA�������ܵ��Ź��ʣ���Ҫ���ݵ�ǰ�Ż�״̬���������Ͳ��ԡ�
        WOA���Ĳ���˵����
        - ��������a������̽����a>1���뿪����a<1������ͳ�����½�����2��0��
        - ����ϵ��b�������������µ�ǿ�ȣ�ͨ���̶�Ϊ1��
        - �������ԣ���Χ��ʳ���ֲ����������������£�ƽ�⣩�����������ȫ��̽����
        
        �Ż�Ŀ�꣺�ڱ�����Ⱥ�����Ե�ͬʱ��������������ֲ����š�
        
        ��ǰ״̬��
        {state}
        
        �������
        1. �����aֵ����ʽ��a=������ֵ����Χ0~2��
        2. �����bֵ����ʽ��b=������ֵ����Χ0.5~2��
        3. ����Ĳ��Ը��ʣ���ʽ��̽������=P1����������=P2������P1+P2=1��
        ���ʾ����
        a=1.2
        b=1.0
        ̽������=0.6����������=0.4
        """
        
        # ����LLM��ȡ����
        response = self.llm.generate_content(prompt)
        return self._parse_response(response.text)
    
    def _parse_response(self, text):
        """����LLM�������ȡ�����Ͳ���"""
        try:
            a = float(re.findall(r'a=([\d.]+)', text)[0])
            b = float(re.findall(r'b=([\d.]+)', text)[0])
            explore_p = float(re.findall(r'̽������=([\d.]+)', text)[0])
            exploit_p = float(re.findall(r'��������=([\d.]+)', text)[0])
            return {
                'a': np.clip(a, 0, 2),
                'b': np.clip(b, 0.5, 2),
                'explore_prob': np.clip(explore_p, 0, 1),
                'exploit_prob': np.clip(exploit_p, 0, 1)
            }
        except:
            # ����ʧ��ʱ����Ĭ��ֵ
            return {'a': 1.0, 'b': 1.0, 'explore_prob': 0.5, 'exploit_prob': 0.5}
    
    def get_escape_strategy(self, current_best):
        """���㷨����ֲ�����ʱ����ȡLLM���������Խ���"""
        prompt = f"""��ǰWOA����ֲ����ţ�������Ӧ��������������ԸĽ�����ǰֵ��{current_best:.6f}����
        ������һ����Ⱥ�Ŷ����ԣ������㷨�����ֲ����š�
        �����ʽ��
        �Ŷ�����=P��0~0.3����Ҫ�Ŷ��ĸ��������
        ����ǿ��=S��0~0.2������������ռ䷶Χ������������
        ʾ����
        �Ŷ�����=0.2
        ����ǿ��=0.05
        """
        
        response = self.llm.generate_content(prompt)
        return self._parse_escape_response(response.text)
    
    def _parse_escape_response(self, text):
        """�����������Ե���Ӧ"""
        try:
            perturb_ratio = float(re.findall(r'�Ŷ�����=([\d.]+)', text)[0])
            noise_strength = float(re.findall(r'����ǿ��=([\d.]+)', text)[0])
            return {
                'perturb_ratio': np.clip(perturb_ratio, 0, 0.3),
                'noise_strength': np.clip(noise_strength, 0, 0.2)
            }
        except:
            # ����ʧ��ʱ����Ĭ��ֵ
            return {'perturb_ratio': 0.2, 'noise_strength': 0.05}

# ������������ȡ�Ż�״̬
def get_optimization_state(pop, fitness, historical_best, iter, max_iter):
    """��ȡWOA��ǰ�Ż�״̬��ת��ΪLLM�������ı�����"""
    # ������Ⱥ�����ԣ����������Ž��ƽ�����룩
    best_idx = np.argmin(fitness)
    diversity = np.mean([np.linalg.norm(ind - pop[best_idx]) for ind in pop])
    
    # ������ڸĽ��������5����
    recent_improve = np.mean(np.diff(historical_best[-5:])) if len(historical_best)>=5 else 0
    
    state = f"""��ǰ�Ż�״̬��
    - �������ȣ�{iter}/{max_iter}��{iter/max_iter*100:.1f}%��
    - ������Ӧ�ȣ�{np.min(fitness):.6f}
    - ��Ⱥ�����ԣ�{diversity:.6f}��ֵԽ�������Խ�ߣ�
    - ���ڸĽ�����{recent_improve:.6f}��ֵԽ��Ľ�Խ�죩
    - ����״̬��{'����̽��' if iter/max_iter < 0.3 else '���ڿ���' if iter/max_iter < 0.7 else '��������'}
    """
    return state

# ��������������Ƿ�����ֲ�����
def check_stuck(historical_best, patience=10, epsilon=1e-6):
    """�ж��Ƿ�����ֲ����ţ�����patience���������Ľ���"""
    if len(historical_best) < patience:
        return False
    recent = historical_best[-patience:]
    improvement = recent[0] - recent[-1]
    return improvement < epsilon  # �Ľ���С����ֵ���ж�Ϊͣ��

# ��ͳWOA�㷨
def traditional_woa(func, dim, pop_size=30, max_iter=100):
    """��ͳ�����Ż��㷨"""
    # ��ʼ����Ⱥ
    lb, ub = func.lb, func.ub
    pop = np.random.uniform(lb, ub, (pop_size, dim))
    fitness = np.array([func.evaluate(ind) for ind in pop])
    best_idx = np.argmin(fitness)
    best_solution = pop[best_idx].copy()
    best_fitness = [fitness[best_idx]]  # ��¼��ʷ����
    
    for iter in range(max_iter):
        # ��ͳWOA������a�����½���2��0
        a = 2 - iter * (2 / max_iter)
        b = 1  # �̶�����ϵ��
        
        for i in range(pop_size):
            # ���ѡ��һ��������ΪĿ��
            r1 = np.random.rand()  # [0,1)�����
            r2 = np.random.rand()  # [0,1)�����
            
            A = 2 * a * r1 - a  # ϵ������
            C = 2 * r2          # ϵ������
            
            p = np.random.rand()
            if p < 0.5:
                if np.abs(A) >= 1:
                    # �������
                    rand_idx = np.random.randint(pop_size)
                    X_rand = pop[rand_idx]
                    D = np.abs(C * X_rand - pop[i])
                    pop[i] = X_rand - A * D
                else:
                    # ��Χ��ʳ
                    D = np.abs(C * best_solution - pop[i])
                    pop[i] = best_solution - A * D
            else:
                # ��������
                D = np.abs(best_solution - pop[i])
                l = (np.random.rand() - 1) * 2  # [-1,1)�����
                pop[i] = D * np.exp(b * l) * np.cos(2 * np.pi * l) + best_solution
            
            # �߽紦��
            pop[i] = np.clip(pop[i], lb, ub)
        
        # ��������Ⱥ���������Ž�
        fitness = np.array([func.evaluate(ind) for ind in pop])
        current_best_idx = np.argmin(fitness)
        if fitness[current_best_idx] < best_fitness[-1]:
            best_solution = pop[current_best_idx].copy()
        best_fitness.append(np.min(fitness))
    
    return best_solution, best_fitness

# LLM-WOA�㷨
def llm_woa(func, dim, pop_size=30, max_iter=100, llm_advisor=None):
    """���LLM�ĸĽ�WOA�㷨"""
    if llm_advisor is None:
        raise ValueError("�����ṩLLM����ʵ��")
    
    # ��ʼ����Ⱥ
    lb, ub = func.lb, func.ub
    pop = np.random.uniform(lb, ub, (pop_size, dim))
    fitness = np.array([func.evaluate(ind) for ind in pop])
    best_idx = np.argmin(fitness)
    best_solution = pop[best_idx].copy()
    best_fitness = [fitness[best_idx]]  # ��¼��ʷ����
    
    # LLM����Ƶ�ʣ�ÿ5������һ�Σ�ƽ��Ч�������ܣ�
    llm_call_interval = 5
    llm_strategy = {'a': 2.0, 'b': 1.0, 'explore_prob': 0.5, 'exploit_prob': 0.5}
    
    for iter in range(max_iter):
        # ÿ���һ������������LLM���²���
        if iter % llm_call_interval == 0:
            state = get_optimization_state(pop, fitness, best_fitness, iter, max_iter)
            llm_strategy = llm_advisor.get_woa_strategy(state)
        
        a = llm_strategy['a']  # ��LLM��̬��������������
        b = llm_strategy['b']  # ��LLM��̬����������ϵ��
        explore_prob = llm_strategy['explore_prob']
        
        # ����Ƿ�����ֲ����ţ�����������LLM��ȡ��������
        if check_stuck(best_fitness):
            escape_strategy = llm_advisor.get_escape_strategy(best_fitness[-1])
            # ִ���Ŷ�
            perturb_ratio = escape_strategy['perturb_ratio']
            noise_strength = escape_strategy['noise_strength']
            perturb_idx = np.random.choice(pop_size, int(pop_size*perturb_ratio), replace=False)
            pop[perturb_idx] += noise_strength * (ub - lb) * np.random.normal(0, 1, (len(perturb_idx), dim))
            pop[perturb_idx] = np.clip(pop[perturb_idx], lb, ub)
        
        for i in range(pop_size):
            # ���ѡ��һ��������ΪĿ��
            r1 = np.random.rand()  # [0,1)�����
            r2 = np.random.rand()  # [0,1)�����
            
            A = 2 * a * r1 - a  # ϵ������
            C = 2 * r2          # ϵ������
            
            # ����LLM����ĸ���ѡ��̽��/��������
            if np.random.rand() < explore_prob:
                # ̽�����ԣ����������ȫ�ֿ�̽��
                if np.abs(A) >= 1:
                    rand_idx = np.random.randint(pop_size)
                    X_rand = pop[rand_idx]
                    D = np.abs(C * X_rand - pop[i])
                    pop[i] = X_rand - A * D
                else:
                    # ��Χ��ʳ���ֲ�������
                    D = np.abs(C * best_solution - pop[i])
                    pop[i] = best_solution - A * D
            else:
                # �������ԣ��������£���ϸ������
                D = np.abs(best_solution - pop[i])
                l = (np.random.rand() - 1) * 2  # [-1,1)�����
                pop[i] = D * np.exp(b * l) * np.cos(2 * np.pi * l) + best_solution
            
            # �߽紦��
            pop[i] = np.clip(pop[i], lb, ub)
        
        # ��������Ⱥ���������Ž�
        fitness = np.array([func.evaluate(ind) for ind in pop])
        current_best_idx = np.argmin(fitness)
        if fitness[current_best_idx] < best_fitness[-1]:
            best_solution = pop[current_best_idx].copy()
        best_fitness.append(np.min(fitness))
    
    return best_solution, best_fitness

# ����ʵ�鲢�ȽϽ��
def run_experiment(dim=10, pop_size=30, max_iter=100, trials=5, save_dir="./results"):
    """���д�ͳWOA��LLM-WOA�ĶԱ�ʵ��"""
    # ������������Ŀ¼
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    if not os.path.exists(os.path.join(save_dir, "convergence_plots")):
        os.makedirs(os.path.join(save_dir, "convergence_plots"))
    
    # ��ʼ��DeepSeekģ�ͺ�LLM����
    api_key = "�����������API"
    deepseek_model = DeepSeekModel(api_key=api_key)
    llm_advisor = LLM_WOA_Advisor(deepseek_model)
    
    # ѡ��CEC2022���Ժ���������ѡ��ǰ5����Ϊʾ�����ɸ�����Ҫ��չ��
    cec_functions = [
        F12022(dim), F22022(dim), F32022(dim),
        F42022(dim), F52022(dim)
    ]
    func_names = [f"F{i+1}2022" for i in range(len(cec_functions))]
    
    # �洢������б�
    results = []
    
    for func_idx, func in enumerate(cec_functions):
        func_name = func_names[func_idx]
        print(f"�������� {func_name} (ά��: {dim})...")
        
        # �洢�������Ľ��
        traditional_results = []
        llm_woa_results = []
        traditional_traces = []
        llm_woa_traces = []
        
        for trial in range(trials):
            print(f"  ���� {trial+1}/{trials}")
            
            # ����������ӣ���֤�Աȹ�ƽ��
            np.random.seed(42 + trial)
            
            # ���д�ͳWOA
            start_time = time.time()
            best_sol_traditional, trace_traditional = traditional_woa(
                func, dim, pop_size, max_iter
            )
            traditional_time = time.time() - start_time
            traditional_best = np.min(trace_traditional)
            traditional_results.append({
                'best': traditional_best,
                'time': traditional_time
            })
            traditional_traces.append(trace_traditional)
            
            # ����LLM-WOA
            start_time = time.time()
            best_sol_llm, trace_llm = llm_woa(
                func, dim, pop_size, max_iter, llm_advisor
            )
            llm_time = time.time() - start_time
            llm_best = np.min(trace_llm)
            llm_woa_results.append({
                'best': llm_best,
                'time': llm_time
            })
            llm_woa_traces.append(trace_llm)
        
        # ����ͳ�ƽ��
        traditional_best_vals = [res['best'] for res in traditional_results]
        traditional_times = [res['time'] for res in traditional_results]
        
        llm_best_vals = [res['best'] for res in llm_woa_results]
        llm_times = [res['time'] for res in llm_woa_results]
        
        # ������
        results.append({
            'function': func_name,
            'traditional_best': np.min(traditional_best_vals),
            'traditional_mean': np.mean(traditional_best_vals),
            'traditional_std': np.std(traditional_best_vals),
            'traditional_time': np.mean(traditional_times),
            'llm_woa_best': np.min(llm_best_vals),
            'llm_woa_mean': np.mean(llm_best_vals),
            'llm_woa_std': np.std(llm_best_vals),
            'llm_woa_time': np.mean(llm_times)
        })
        
        # ������������
        plt.figure(figsize=(10, 6))
        
        # ��ͳWOAƽ����������
        traditional_mean_trace = np.mean(traditional_traces, axis=0)
        plt.plot(traditional_mean_trace, 'b-', linewidth=2, label='��ͳWOA')
        
        # LLM-WOAƽ����������
        llm_mean_trace = np.mean(llm_woa_traces, axis=0)
        plt.plot(llm_mean_trace, 'r-', linewidth=2, label='LLM-WOA')
        
        plt.xlabel('��������')
        plt.ylabel('������Ӧ��ֵ')
        plt.title(f'{func_name} �������� (ά��: {dim})')
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.legend()
        plt.yscale('log')  # ʹ�ö����߶ȸ��ʺ�չʾ�Ż�����
        plt.tight_layout()
        
        # ������������
        plot_path = os.path.join(save_dir, "convergence_plots", f"{func_name}_dim{dim}.png")
        plt.savefig(plot_path, dpi=300)
        plt.close()
        print(f"  ���������ѱ�����: {plot_path}")
    
    # ��������浽Excel
    df = pd.DataFrame(results)
    excel_path = os.path.join(save_dir, f"woa_vs_llm_woa_dim{dim}.xlsx")
    
    # ����Excel�����������ø�ʽ
    wb = Workbook()
    ws = wb.active
    ws.title = f"ά��_{dim}"
    
    # �������
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # ���ñ�ͷ��ʽ
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    
    # �����ڴ�ͳWOA����ֵ���мӺڴ���
    bold_font = Font(bold=True)
    # �ӵڶ��п�ʼ��������ͷ���������һ��
    for row in range(2, ws.max_row + 1):
        # �Ƚ�����ֵ��ֵԽСԽ�ã�
        if ws.cell(row=row, column=6).value < ws.cell(row=row, column=2).value:
            ws.cell(row=row, column=6).font = bold_font
        
        # �Ƚ�ƽ��ֵ��ֵԽСԽ�ã�
        if ws.cell(row=row, column=7).value < ws.cell(row=row, column=3).value:
            ws.cell(row=row, column=7).font = bold_font
        
        # �Ƚϱ�׼�ֵԽСԽ�ã�
        if ws.cell(row=row, column=8).value < ws.cell(row=row, column=4).value:
            ws.cell(row=row, column=8).font = bold_font
    
    # ����Excel�ļ�
    wb.save(excel_path)
    print(f"ʵ�����ѱ�����: {excel_path}")
    
    return df

if __name__ == "__main__":
    # ����ʵ��
    # ���Ե���������ά��(dim)����Ⱥ��С(pop_size)������������(max_iter)���������(trials)
    run_experiment(
        dim=10,          # ����ά��
        pop_size=30,     # ��Ⱥ��С
        max_iter=100,    # ����������
        trials=5,        # ÿ�β��Ժ������е��������
        save_dir="./woa_comparison_results"  # �������Ŀ¼
    )
    print("����ʵ����ɣ�")
